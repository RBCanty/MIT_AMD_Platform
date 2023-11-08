# To supplant how the Spark is controlled and data is managed on the AMD platform
# Ben Canty - 5/31/2022
import datetime
import json
import os
import re
import time
from copy import copy
from functools import wraps
from os import path, listdir, getcwd, getpid
from threading import Lock
from typing import Tuple, Optional, Any, List, Union

import openpyxl as pyxl

from constants import WELL_IDS, get_flipped_well_id
from custom_classes import Narg, safe_open, grab_newest_folder
from database_constants import DBG_CONTAINER_NAME
from functionals import list_subtract

_METHOD_ABS_SCAN = r".\Spark_Templates\MCN_AbsScan.xml"
_METHOD_PL_SCAN = r".\Spark_Templates\MCN_PLScan.xml"
_METHOD_PREHEAT = r".\Spark_Templates\MCN_Preheat.xml"
try:
    open(_METHOD_ABS_SCAN, 'r')
except FileNotFoundError:
    _METHOD_ABS_SCAN = r".\Spark_API\Spark_Templates\MCN_AbsScan.xml"
    _METHOD_PL_SCAN = r".\Spark_API\Spark_Templates\MCN_PLScan.xml"
    _METHOD_PREHEAT = r".\Spark_API\Spark_Templates\MCN_Preheat.xml"
try:
    open(_METHOD_ABS_SCAN, 'r')
except FileNotFoundError as fnfe:
    raise OSError(f"Spark Method Templates could not be found in '.\\(Spark_API)\\Spark_Templates\\' "
                  f"from current working directory: '{getcwd()}'") from fnfe

_DEFAULT_MCN_DIR = r"C:\Users\kfj_AMD_FTIR\Documents\MCN_Spark_Data"
_DEFAULT_SAVE_DIR = r"C:\SparkData"

ABS_TEMPLATE_MAP = {
    'start': 'ABSST_WL_LB',
    'measure_count': 'ABSST_MC',
    'settle': 'ABSST_ST',  # May not do anything
    'step': 'ABSST_STEP_SIZE',
    'stop': 'ABSST_WL_UB',
}
PL_TEMPLATE_MAP = {
    'start': 'PLST_WL_LB',
    'stop': 'PLST_WL_UB',
    'step': 'PLST_STEP_SIZE',
    'excite': 'PLST_XCITE',
}
THERMAL_TEMPLATE_MAP = {
    't_control': 'THST_T_CTRL',
    'temperature': 'THST_SETPOINT',
    'max_t': 'THST_UB',
    'min_t': 'THST_LB',
    'off_when_done': 'THST_PMTO',
    'wait_for_t': 'THST_WAIT'
}

_HEADER_LIST = ["Date", "Time", "System", "User", "Plate", "Lid lifter", "Humidity Cassette",
                "Smooth mode", "Plate area", "Mode", "Name", "Wavelength start [nm]",
                "Wavelength end [nm]", "Wavelength step size [nm]", "Number of flashes",
                "Settle time [ms]", "Part of Plate", "Start Time", "Temperature [°C]", "End Time",
                "Warning", "Error"]

_FIND_WELL_REGEX = r'AlphanumericCoordinate="(\w*)" Column="(\w*)" Grid="(\w*)" ' \
                   r'IdentifierGroup="None" IsFlagged="False" IsOut="False" IsSelected="True"'
SPARK_METHOD_KEYWORDS = ['mode',
                         'start', 'stop', 'step', 'settle', 'excite',
                         'iterations', 'interval', 'is_photo',
                         'temperature', 'tol', 'off_when_done', 'wait_for_t']

TIME_FORMAT = "%Y-%m-%d--%H:%M:%S"


class SparkMethod:
    """ SparkMethod (class: dictionary of method specifications)

    Class for storing and vetting method details and creating xml scripts for the Spark instrument to execute.
      * hist: (internal) used to log events where inputs are adjusted to fit allowable bounds
      * mode: Mode of instrument (abs, pl, ...) to guide which properties should be used
      * start: Lower scan value (spark units or nm -- self corrects)
      * stop: Upper scan value (spark units or nm -- self corrects)
      * step: Scan step size (spark units or nm -- self corrects)
      * settle: Settling time (may be defunct, hard to notice the instrument's behavior change)
      * excite: (for pl) Excitation wavelength (spark units or nm -- self corrects)
      * iterations: Number of times the plate will be scanned
      * interval: (is_photo=False) time to wait between scans in minutes, (is_photo=True) exposure time under light
      * is_photo: If the interval is a wait time or exposure time (should the solar reactor be used)
      * offset: Starts the iteration number at a value other than 0 (used when multiple scans within a single queue are
        needed that require liquid handling between them, e.g. hyperpolarizability)
      * temperature: Target temperature (-1 to deactivate temperature control)
      * photo_temp: Target temperature for Photoreactor (-1 to deactivate temperature control)
      * tol: The tolerance of the temperature set-point (<number>: +/- tolerance, <list>: lower & upper bounds)
      * off_when_done: Boolean if the temperature control should be turned off after the run
      * wait_for_t: Boolean if the Spark should wait for the temperature to reach the setpoint before starting the scan
    """
    def __init__(self, **kwargs):
        self.history = kwargs.get('hist', list())
        self.mode = kwargs.get('mode', 'abs')

        # Spectral Properties
        self._start = int(kwargs.get('start', 3000))
        self._stop = int(kwargs.get('stop', 9000))
        self._step = int(kwargs.get('step', 20))
        self._settle = self.bound(int(kwargs.get('settle', 50)), 50, 1000)
        self._excite = kwargs.get('excite', Narg())

        # Loop Properties
        self._n_iterations = int(kwargs.get('iterations', 1))
        self._interval = kwargs.get('interval', -1)
        self._photo = self.decode_bool(kwargs.get('is_photo', False))
        self._offset = int(kwargs.get('offset', 0))

        # Temperature Properties
        self._setpoint = float(kwargs.get('temperature', -1))
        self._tolerance = kwargs.get('tol', 0.5)
        self._off = self.decode_bool(kwargs.get('off_when_done', True))
        self._wait_for_t = self.decode_bool(kwargs.get('wait_for_t', False))

        # Other Properties
        self._photo_temp = self.bound(float(kwargs.get('photo_temp', -1)), None, 60, float)

    def __dict__(self):
        representation = {'hist': self.history,
                          'mode': self.mode,
                          'start': self._start,
                          'stop': self._stop,
                          'step': self._step,
                          'settle': self._settle,
                          'iterations': self._n_iterations,
                          'interval': self._interval,
                          'offset': self._offset,
                          'is_photo': self._photo,
                          'temperature': self._setpoint,
                          'tol': self._tolerance,
                          'off_when_done': self._off,
                          'wait_for_t': self._wait_for_t,
                          'photo_temp': self._photo_temp}
        if not isinstance(self._excite, Narg):
            representation.update({'excite', self._excite})
        return representation

    def bound(self, var, lower, upper, dtype: Optional[Any] = None):
        """
        Converts a number to a value within a specified domain.
        (e.g. 4 on [5,8) becomes 5)

        Will log any overwriting of the value to the 'self.history' field.

        :param var: The value being bounded
        :param lower: The lower bound for the number
        :param upper: The upper bound for the number
        :param dtype: The desired Type of the result [None: dtype = type(var)]
        :return: A corrected version of the number
        """
        if dtype is None:
            dtype = type(var)
        if (lower is not None) and (var < lower):
            self.history.append(f"Warning: {var} < lower bound ({lower}), {var} -> {lower}")
            return dtype(lower)
        if (upper is not None) and (var > upper):
            self.history.append(f"Warning: {var} > upper bound ({upper}), {var} -> {upper}")
            return dtype(upper)
        return dtype(var)

    def decode_bool(self, boolean_repr: Tuple[bool, str, int, float]) -> bool:
        """
        Helper method to convert boolean representations to a boolean.
          * True, 't', 'true', 'yes', 'y', '1', and <not zero> map to True (case-insensitive)
          * False, 'f', 'false', 'no', 'n', '0', and <zero> map to False (case-insensitive)

        :param boolean_repr: A boolean, string, or number representation of a boolean
        :return: The boolean value of the representation

        :raises ValueError: If the representation is not recognized
        """
        if isinstance(boolean_repr, bool):
            return boolean_repr
        if isinstance(boolean_repr, str):
            if boolean_repr.lower() in ['t', 'true', 'yes', 'y', '1']:
                self.history.append(f"Boolean representation '{boolean_repr} (str)' -> True")
                return True
            if boolean_repr.lower() in ['f', 'false', 'no', 'n', '0']:
                self.history.append(f"Boolean representation '{boolean_repr} (str)' -> False")
                return False
            raise ValueError(f"Boolean representation '{boolean_repr}' not recognized")
        if isinstance(boolean_repr, (int, float)):
            self.history.append(f"Boolean representation '{boolean_repr} ({type(boolean_repr)})'"
                                f" -> {bool(boolean_repr)}")
            return bool(boolean_repr)
        raise ValueError(f"Boolean representation '{boolean_repr}' not recognized")

    def convert_units(self, val, scale, break_point):
        """
        Helper method to convert specifications (either Spark units or nm) into Spark units.  This only works because
          the allowable ranges expressed in each unit system do not overlap).  Logs updates.

        :param val: The value being converted
        :param scale: The conversion factor (nm)*(scale) = (Spark units)
        :param break_point: The boundary for the two unit domains.
        :return: The value of val expressed in Spark units.
        """
        if val < break_point:
            self.history.append(f"Value given in nm ({val}), converting to Spark Units: {val} -> {scale*val}")
            return scale * val
        return val

    @staticmethod
    def sanity_check(start, stop, step, count):
        """
        Helper method to check start-step-stop values.

        :param start: The lower bound
        :param stop: The upper bound
        :param step: The step size
        :param count: The number of measurements
        :return: None

        :raises ValueError: If values do not make sense ('stop < start'; 'step > stop - start'; or 'count < 1')
        """
        if stop < start:
            raise ValueError(f"Stop value ({stop / 10.0} nm) is less than Start value ({start / 10.0} nm)")
        if step > stop - start:
            raise ValueError(f"Step size ({step / 10.0} nm) is greater than measurement space "
                             f"({start / 10.0} - {stop / 10.0} nm)")
        if count < 1:
            raise ValueError(f"No measurements will be taken")
        return None

    def get_abs_properties(self) -> dict:
        """
        Used to get usable abs-mode spectral properties from those specified.

        :return: A dictionary containing the corrected values for start, stop, step, and measure_count.

        :raises ValueError: from 'self.sanity_check' method if specified values cannot be rectified
        """
        start = self.convert_units(self._start, 10, 2000)
        start = self.bound(start, 2000, 10000, int)
        stop = self.convert_units(self._stop, 10, 2000)
        stop = self.bound(stop, 2000, 10000, int)
        if start > stop:
            self.history.append(f"Start ({start}) is greater than Stop ({stop}), swapping values")
            temp = start
            start = stop
            stop = temp
        step = self.convert_units(self._step, 10, 20)
        step = self.bound(step, 20, 500, int)

        measure_count = int(1 + (stop - start) / step)
        stop = (measure_count - 1) * step + start

        self.sanity_check(start, stop, step, measure_count)

        return {'start': start,
                'step': step,
                'stop': stop,
                'measure_count': measure_count}

    def get_pl_properties(self) -> dict:
        """
        Used to get usable pl-mode spectral properties from those specified.

        :return: A dictionary containing the corrected values for start, stop, step, measure_count, and excite.

        :raises ValueError: from 'self.sanity_check' method if specified values cannot be rectified, or if the excite
          parameter is either missing or cannot be parsed
        """
        start = self.convert_units(self._start, 10, 2000)
        start = self.bound(start, 2800, 9000, int)
        stop = self.convert_units(self._stop, 10, 2000)
        stop = self.bound(stop, 2800, 9000, int)
        if start > stop:
            self.history.append(f"Start ({start}) is greater than Stop ({stop}), swapping values")
            temp = start
            start = stop
            stop = temp
        step = self.convert_units(self._step, 10, 20)
        step = self.bound(step, 20, 500, int)
        if not isinstance(self._excite, Narg):
            try:
                self._excite = int(self._excite)
            except (ValueError, TypeError):
                raise ValueError(f"Excitation parameter '{self._excite}' could not be parsed")
        else:
            raise ValueError(f"Excitation parameter not specified")
        excite = self.convert_units(self._excite, 10, 2000)
        excite = self.bound(excite, 2300, min(start - 500, 9000), int)

        measure_count = int(1 + (stop - start) / step)
        stop = (measure_count - 1) * step + start

        self.sanity_check(start, stop, step, measure_count)

        return {'start': start,
                'step': step,
                'stop': stop,
                'measure_count': measure_count,
                'excite': excite}

    def get_loop_properties(self) -> dict:
        """
        Used to get usable loop properties from those specified.

        :return: A dictionary containing the corrected values for n_iterations, interval, is_photo, and offset.
        """
        n_iterations = self.bound(self._n_iterations, 1, 999, int)
        if self._interval is None:
            interval = None
        elif self._interval <= 0:
            interval = None
        else:
            interval = float(self._interval)
        return {'n_iterations': n_iterations,
                'interval': interval,
                'is_photo': self._photo,
                'offset': self._offset}

    def get_spectral_properties(self) -> dict:
        """
        Used to get usable spectral properties from those specified.

        :return: A dictionary containing the corrected values.

        :raises ValueError: from 'self.get_abs_properties' or 'self.get_pl_properties' methods, or if 'self.mode' is not
          recognized
        """
        # Spectral
        if self.mode == 'abs':
            return self.get_abs_properties()
        elif self.mode == 'pl':
            return self.get_pl_properties()
        elif self.mode == 'preheat':
            return {}
        else:
            raise ValueError(f"Mode '{self.mode}' not recognized")

    def get_temperature_properties(self):
        """
        Used to get usable temperature properties from those specified.

        :return: A dictionary containing the corrected values for temperature, max_t, min_t, off_when_done,
          wait_for_t, and t_control.

        :raises ValueError: if the tolerance parameter is not a number or list, or if a specification is missing.
        """
        tp = dict()
        if self._setpoint < 0:
            return {'temperature': "20.0",
                    'max_t': "25.0",
                    'min_t': "20.0",
                    'off_when_done': "True",
                    'wait_for_t': "False",
                    't_control': "Off"}
        else:
            tp['t_control'] = "On"

        if self._photo:
            setpoint = 42.0
        else:
            setpoint = self.bound(self._setpoint, 20, 42, float)

        tp['temperature'] = "{:.1f}".format(setpoint)
        if isinstance(self._tolerance, (tuple, list)):
            min_t, max_t = self._tolerance
            min_t = self.bound(min_t, 20, 42, float)
            max_t = self.bound(max_t, 20, 42, float)
            if min_t > max_t:
                self.history.append(f"Min Temperature ({min_t}) is greater than Maximum Temperature ({max_t}),"
                                    f" swapping values")
                temp = min_t
                min_t = max_t
                max_t = temp
            tp['min_t'] = "{:.1f}".format(min_t)
            tp['max_t'] = "{:.1f}".format(max_t)
        elif isinstance(self._tolerance, (int, float)):
            tolerance = self.bound(self._tolerance, 0.5, 22, float)
            tp['min_t'] = "{:.1f}".format(self.bound(setpoint - tolerance, 20, 42))
            tp['max_t'] = "{:.1f}".format(self.bound(setpoint + tolerance, 20, 42))
        else:
            raise ValueError(f"Temperature tolerance must be a number or list")
        tp['off_when_done'] = str(self._off)
        tp['wait_for_t'] = str(self._wait_for_t)

        problems = list()
        for validation_key in ['t_control', 'temperature', 'max_t', 'min_t',
                               'off_when_done', 'wait_for_t', 't_control']:
            if validation_key not in tp:
                problems.append(f"{validation_key}")
        if problems:
            raise ValueError(f"Error: Temperature property specifications missing: {', '.join(problems)}")

        return tp

    def get_photo_temperature(self):
        return self._photo_temp

    def generate_method_file(self, override=None, excluded_wells=None, transient=None) -> str:
        """
        Used to create the xml string used by the Spark to execute a method.

        :param override: Allows a mode to be specified regardless of the value stored in the 'self.mode' property
        :param excluded_wells: A list of wells to not scan
        :param transient: Is this a final scan or transient scan
        :return: A string representation of the xml file used by the Spark

        :raises ValueError: If excitation parameter missing from a pl method, if mode is not recognized, from
          'self.get_abs_properties' or 'self.get_pl_properties' as determined by the mode, and from
          'self.get_temperature_properties'.
        """
        total_mapping = dict()
        total_specifications = dict()
        # Select mode
        mode = override if override else self.mode
        if mode == 'pl' and isinstance(self._excite, Narg):
            raise ValueError("A pl method must have the excitation wavelength defined")

        # Spectral
        if mode == 'abs':
            total_specifications.update(self.get_abs_properties())
            total_mapping.update(ABS_TEMPLATE_MAP)
            with open(_METHOD_ABS_SCAN, 'r') as fh:
                method_core = fh.read()
        elif mode == 'pl':
            total_specifications.update(self.get_pl_properties())
            total_mapping.update(PL_TEMPLATE_MAP)
            with open(_METHOD_PL_SCAN, 'r') as fh:
                method_core = fh.read()
        elif mode == 'preheat':
            with open(_METHOD_PREHEAT, 'r') as fh:
                method_core = fh.read()
        else:
            raise ValueError(f"Mode '{mode}' not recognized")

        # Temperature
        total_specifications.update(self.get_temperature_properties())
        if transient:
            total_specifications['off_when_done'] = "False"
        total_mapping.update(THERMAL_TEMPLATE_MAP)

        # Settle Time
        total_specifications.update({'settle': self.bound(self._settle, 50, 1000)})

        # Fill Out Template form
        for k, v in total_specifications.items():
            try:
                method_core = method_core.replace(total_mapping[k], str(v))
            except KeyError:
                pass

        # Exclude Wells
        method_core = self.deselect_wells(method_core, excluded_wells)

        return method_core

    @staticmethod
    def deselect_wells(method_file_str, list_of_wells, log: list = None) -> str:
        """
        Removes this list of wells from the method file

        :param log: A list used to log problems
        :param method_file_str: The string representation of the method
        :param list_of_wells: A list of wells in the form ["A1", "B3", "D12", ...].
          "A:" or ":5" will remove the entire row (A) or column (5), respectively
        :return: method_file or the original version if error
        """
        if log is None:
            log = list()
        if list_of_wells is None:
            return method_file_str

        original_method_file_str = copy(method_file_str)
        if any([":" in well for well in list_of_wells]):
            new_wells = []
            og_well_list = copy(list_of_wells)
            for well in og_well_list:
                if ":" in well:
                    well_core: str = well.replace(":", "")
                    if well_core.isnumeric():
                        new_wells += [f"{L}{well_core}" for L in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']]
                    elif well_core.isalpha():
                        new_wells += [f"{well_core}{N+1}" for N in range(0, 12)]
                    list_of_wells.__delitem__(list_of_wells.index(well))
            list_of_wells += new_wells

        # Clean up possible duplicate entries
        temp = list()
        [temp.append(x) for x in list_of_wells if x not in temp]
        list_of_wells = temp

        # All methods calling deselect_wells are using the Database well ID (the True well ID)
        # However, due to RoMa problems, the plate is loaded revered (H12 is where A1 should be)
        # So the Spark
        list_of_wells = [get_flipped_well_id(_well) for _well in list_of_wells]

        match_object = re.finditer(_FIND_WELL_REGEX, method_file_str)
        for match in match_object:
            try:
                old_line, well_id, column, grid = match.group(0), match.group(1), match.group(3), match.group(3)
            except IndexError:
                log.append("Error in XML file, incomplete match found")
                return original_method_file_str
            if well_id in list_of_wells:
                new_line = old_line.replace('IsSelected="True"', 'IsSelected="False"')
                method_file_str = method_file_str.replace(old_line, new_line)
        return method_file_str


class SparkData:
    """ SparkData (class: raw data processor)

    Used to load *.xlsx output files from the Spark and convert to a pythonic form
    """
    def __init__(self, filepath: str):
        """

        :param filepath: The path to the Spark output data file (xlsx)
        """
        self._HEADER_LIST = _HEADER_LIST
        self._COMMENT = ["#", "%", "$"]

        self.filepath = filepath
        self.spark_metadata = dict()
        """
        The non-experimental data stored in the header of the raw file
        """
        self.experiment_data = dict()
        """
        The experimental data:
          * all_well_ids = list of all well IDs specified in the file (order matches signal_values: first level)
          * wavelengths = list of wavelengths measured (order matches signal_values: second level)
          * signal_values = list of lists of all signal values (OVER -> None) for each well
        """

    def _isheader(self, x: str):
        """
        Checks if a string is a recognized header, and if so, which type.

        :param x: A string header
        :return: "H" (base header), "W" (well_id, heads a list of abs values),
          "T" (table header, heads a list of wavelengths), "C" (comment), False - (not header)
        """
        is_named = x in self._HEADER_LIST
        is_well = x in WELL_IDS
        is_table_header = x == "Wavel."
        is_comment = x in self._COMMENT
        if is_named:
            return "H"
        elif is_well:
            return "W"
        elif is_table_header:
            return "T"
        elif is_comment:
            return "C"
        else:
            return False

    def compile(self):
        """
        Opens an xlsx file and saves it to the 'self.spark_metadata' and 'self.experiment_data' properties

        Dictionary is keyed by the headers of the file and 3 additional keys (WELL_IDS, WAVELENGTHS, ABS_VALUES) that
        store the X axis values, Y axis values, and Z values.

        Opens a file via os.path.join(dir_path, filename) and openpyxl::load_workbook

        :return: A list of booleans for if each well ID was completed
        """
        self.spark_metadata["Error Flag"] = False
        for k in self._HEADER_LIST:
            self.spark_metadata.setdefault(k, None)
        wavelengths = list()
        absorbance_values = list()
        well_ids = list()

        workbook = pyxl.load_workbook(self.filepath)
        worksheet = workbook["Result sheet"]

        for row in worksheet.iter_rows():
            active_row = [cell.value for cell in row if cell.value is not None]
            active_length = len(active_row)
            if active_length > 1:
                if isinstance(active_row[0], str):
                    active_row[0] = active_row[0].replace(":", "")
                data_type = self._isheader(active_row[0])
                header = active_row[0]
                values = active_row[1:]

                if ("Warning" in header) or ("Error" in header):
                    self.spark_metadata["Error Flag"] = True
                if data_type == "H":
                    self.spark_metadata[header] = values[0]
                elif data_type == "W":
                    well_ids.append(get_flipped_well_id(header))  # The Spark has things flipped (due to RoMa issues)
                    absorbance_values.append([float(value) if not isinstance(value, str) else None for value in values])
                elif data_type == "T":
                    wavelengths = [float(value) for value in values if value]
                elif data_type == "C":
                    print(f"Comments: {values}")

        self.experiment_data['all_well_ids'] = well_ids
        self.experiment_data['wavelengths'] = wavelengths
        self.experiment_data['signal_values'] = absorbance_values

        expected_length = len(wavelengths)

        return [well_ids[absorbance_values.index(well_abs_val)]
                for well_abs_val in absorbance_values
                if len(well_abs_val) == expected_length]

    def get_update_list(self):
        """
        Helper method.

        :return: List of the headers for values  that should be overwritten on subsequent attempts of the same scan
        """
        return list_subtract(self._HEADER_LIST,
                             ["Time", "End Time", "Start Time", "Temperature [°C]", "Warning", "Error"])

    @staticmethod
    def get_addendum_list():
        """
        Helper method.

        :return: List of headers for values that should be appended to on subsequent attempts of the same scan
        """
        return ["Time", "End Time", "Start Time", "Temperature [°C]", "Warning", "Error"]


class SparkProjectFileHandler:
    """ SparkProjectFileHandler (class: file manager)

    Manages the creation of a project file as well as the accessing and committing of data to said file.
    If a project at the specified location exist, that project will be used; otherwise a project at the specified
    location will be created.

    Uses Lock to prevent concurrent access to the underlying project file.

    A project is a json dictionary representation of :class:`SparkMethod` and :class:`SparkData` objects.
    """
    def __init__(self, filepath: str):
        """
        Loads the path to the project file (does not create yet)

        References: :class:`Appellomancer`

        :param filepath: Path to a json file containing a project (preferably generated from an Appellomancer)
        """
        self.filepath = filepath
        self.directory, _ = os.path.split(self.filepath)
        self.lockfile = path.join(self.directory, "lock.semaphore")

    def via_lock(func):
        """ Wrapper function for the lock/try-foo/finally-unlock workflow

        Implements a timeout of 10 seconds for the lock

        :param func: Wrapped function
        :return: Return of Wrapped function
        :raise TimeoutError: On Timeout
        """
        @wraps(func)
        def wrapper_function(self, *args, **kwargs):
            for _ in range(1000):
                try:
                    with safe_open(self.lockfile, 'x') as lockfile_handle:
                        lockfile_handle.write(str(getpid()))
                        break
                except IOError:
                    time.sleep(0.01)
            else:
                raise TimeoutError(f"Access to '{self.filepath}' by {func.__name__}() timeout (10 seconds) exceeded.")

            try:
                return func(self, *args, **kwargs)
            finally:
                os.remove(self.lockfile)

        return wrapper_function

    def _touch(self, filepath=None):
        """
        Attempts to open a project file.  If it exits, nothing further is done.  If it does not exist, it will create
        the file (as an empty json dictionary)

        :param filepath: (optional) if specified, use this filepath; otherwise, use 'self.filepath' property
        :return: None
        """
        if filepath is None:
            filepath = self.filepath
        try:
            with safe_open(filepath, 'x') as fh:
                json.dump({}, fh)
        except FileExistsError:
            return

    @via_lock
    def set_spark_method_details(self, details: SparkMethod = None):
        """
        Author method for the spark method details used in a project.

        Creates the project if it does not already exist (will return an empty dictionary)

        References: :class:`SparkMethod`

        :param details: (None) Read method details; (SparkMethod) Update method details
        :return: project['method_specifications']
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            project: dict = json.load(fh)
            project.setdefault('method_specifications', dict())
            if details is None:
                return project['method_specifications']
            else:
                project['method_specifications'].update(details.__dict__())
        with open(self.filepath, 'w') as fh:
            json.dump(project, fh)
        return project['method_specifications']

    def get_spark_method_details(self):
        """
        Author method for the spark method details used in a project.

        Creates the project if it does not already exist (will return an empty dictionary)

        References: :class:`SparkMethod`

        :return: project['method_specifications']
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            project: dict = json.load(fh)
        project.setdefault('method_specifications', dict())
        return project['method_specifications']

    @via_lock
    def add_data(self, spark_output: str, iteration: Union[int, str], wellplate_document: dict):
        """
        Method to add data to a project file.

        Creates the project if it does not already exist (will return an empty dictionary)

        References: :class:`SparkData`

        :param spark_output: Filepath to data (filepath argument of SparkData __init__ method)
        :param iteration: The iteration number for the data being added.  Repeated iteration numbers will overwrite only
          wells in the 'completed_wells' output from compiling the SparkData object.
        :param wellplate_document: Document from database_interface.query_document
        :return: <list> the 'completed_wells' output from compiling the SparkData object
        """
        self._touch()
        spark_data = SparkData(spark_output)
        completed_wells = spark_data.compile()
        update_list = spark_data.get_update_list()
        addended_list = spark_data.get_addendum_list()
        iteration = str(iteration)

        with open(self.filepath, 'r') as fh:
            project: dict = json.load(fh)

        project.setdefault('abscissa', dict())
        project['abscissa'].setdefault('metadata', dict())
        project['abscissa'].setdefault('x-axis', dict())
        project.setdefault('history', list())

        meta_data = spark_data.spark_metadata
        well_ids = spark_data.experiment_data.get('all_well_ids', list())
        wavelengths = spark_data.experiment_data.get('wavelengths', list())
        signal_values = spark_data.experiment_data.get('signal_values', list())

        x_axis_key = 1
        for well_id in completed_wells:
            try:
                i = well_ids.index(well_id)
            except ValueError:
                continue
            project.setdefault(well_id, dict())

            if all(wavelengths != x for _, x in project['abscissa']['x-axis'].items()):
                if project['abscissa']['x-axis']:
                    x_axis_key += 1
                project['abscissa']['x-axis'][str(x_axis_key)] = wavelengths

            project[well_id].setdefault(iteration, dict())

            project[well_id][iteration]['x-axis'] = str(x_axis_key)
            project[well_id][iteration]['y-axis'] = signal_values[i]
            project[well_id][iteration].setdefault('well_context', dict())
            project[well_id][iteration]['well_context'] = wellplate_document.get(
                'contents', dict()).get(well_id, dict())
            project[well_id][iteration]['well_context'][DBG_CONTAINER_NAME] = wellplate_document.get(DBG_CONTAINER_NAME,
                                                                                                     '')
            project['abscissa']['metadata'].setdefault(iteration, dict())

            for k in update_list:
                if k in meta_data:
                    if project['abscissa']['metadata'][iteration].get(k, None) is None:
                        project['abscissa']['metadata'][iteration][k] = meta_data[k]
                    elif project['abscissa']['metadata'][iteration].get(k, None) == meta_data[k]:
                        pass
                    else:
                        project['history'].append(f"({well_id}, {iteration}; {k}): "
                                                  f"{project['abscissa']['metadata'][iteration].get(k, None)} "
                                                  f"updated to {meta_data[k]}")
                        project['abscissa']['metadata'][iteration][k] = meta_data[k]

            for k in addended_list:
                project['abscissa']['metadata'][iteration].setdefault(k, list())
                if k in meta_data:
                    project['abscissa']['metadata'][iteration][k].append(meta_data[k])

        with open(self.filepath, 'w') as fh:
            json.dump(project, fh)

        return completed_wells

    def get_datum_by_well(self, well, iteration: Union[int, str] = 0) -> Tuple[dict, dict, dict, list]:
        """
        Getter method for the data from a specific iteration.

        **Iteration Data**
          * x-axis: list of wavelengths
          * y-axis: list of signal values
          * processing: (tbd)
          * well_context:

            - solvents: [ [name, volume in uL, smiles] , ... ]
            - target_analytes: [ [name, volume in uL, smiles] , ... ]
            - total_volume: volume in uL
            - reagents: [ [name, volume in uL, smiles] , ... ]

        **metadata**
          * Headers from :class:`SparkData` ("Date", "Time", "System", "User", "Plate", "Lid lifter",
            "Humidity Cassette", "Smooth mode", "Plate area", "Mode", "Name", "Wavelength start [nm]",
            "Wavelength end [nm]", "Wavelength step size [nm]", "Number of flashes", "Settle time [ms]",
            "Part of Plate", "Start Time", "Temperature [°C]", "End Time", "Warning", and "Error")
          * Keys "Time", "End Time", "Start Time", "Temperature [°C]", "Warning", and "Error" are lists of values

        **Method Specifications**
          * From :class:`SparkMethod`

        :param well: A well ID.
        :param iteration: The iteration being accessed.
        :return: Quadruple of dictionaries: dictionary of the iteration data, dictionary of the spark raw output
          metadata for the given iteration, a dictionary copy of the method specifications, and a list of the
          'x-axis' value replaced by the values for the x-axis (wavelengths) rather than a pointer
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            data = json.load(fh)
        well_data = data.get(well, {})
        iteration = str(iteration)
        iteration_data = well_data.get(iteration, {})
        x_data = data['abscissa']['x-axis'][str(iteration_data['x-axis'])]

        metadata = data['abscissa']['metadata'].get(iteration, None)

        return iteration_data, metadata, data['method_specifications'], x_data

    @via_lock
    def commit_datum_by_well(self, well, iteration: Union[int, str], iteration_data):
        """
        Performs an update on the project on a particular well.  Updates the entire well-iteration pair.

        for well, iteration, data in generator:
          project[well][str(iteration)] = data

        When performing well-by-well operations, it is on the committer to log an information in the 'processing'
        <List[str]> field of the iteration_data.

        :param well: A well ID
        :param iteration: An iteration number
        :param iteration_data: The data replacing the old data
        :return: None
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            project = json.load(fh)

        project[well][str(iteration)] = iteration_data

        with open(self.filepath, 'w') as fh:
            json.dump(project, fh)

    @via_lock
    def commit_data_by_generator(self, generator, log: Union[list, str]):
        """
        Performs an update on the project obeying a generator.  Updates the entire well-iteration pair.

        for well, iteration, data in generator:
          project[well][str(iteration)] = data

        :param generator: A tuple of the form (well, iteration, data)
        :param log: A string or list of string recounting what processing has occurred.
        :return: None
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            project = json.load(fh)

        for well, iteration, data in generator:
            project[well][str(iteration)] = data

        project.setdefault('history', list())
        if isinstance(log, list):
            project['history'].extend(log)
        if isinstance(log, str):
            project['history'].append(log)

        with open(self.filepath, 'w') as fh:
            json.dump(project, fh)

    def get_data_by_well(self, well) -> Tuple[List[Tuple[int, dict, dict, list]], dict]:
        """
        Extension of 'self.get_datum_by_well' method to fetch all data for a well.

        :param well: A well ID.
        :return: list, sorted by iteration, of (iteration#, iteration data, metadata) tuples and a dictionary
          copy of the method specifications
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            data = json.load(fh)
        ret_data = list()
        well_data = data.get(well, dict())
        for ii in sorted(well_data.keys(), key=lambda x: int(x)):
            iteration_data = well_data[ii]
            # from pprint import pprint
            # pprint(data['abscissa']['x-axis'].keys())
            x_data = data['abscissa']['x-axis'][str(iteration_data['x-axis'])]
            metadata = data['abscissa']['metadata'].get(ii, None)
            ret_data.append((ii, iteration_data, metadata, x_data))

        return ret_data, data['method_specifications']

    def get_spark_metadata(self, iteration: Union[int, str] = 0, key=None, _all=False) -> Any:
        """
        Pulls the metadata for a project.

        * Headers from :class:`SparkData` ("Date", "Time", "System", "User", "Plate", "Lid lifter",
          "Humidity Cassette", "Smooth mode", "Plate area", "Mode", "Name", "Wavelength start [nm]",
          "Wavelength end [nm]", "Wavelength step size [nm]", "Number of flashes", "Settle time [ms]",
          "Part of Plate", "Start Time", "Temperature [°C]", "End Time", "Warning", and "Error")
        * Keys "Time", "End Time", "Start Time", "Temperature [°C]", "Warning", and "Error" are lists of values

        :param iteration: The iteration # for the requested metadata
        :param key: Requests a specific item from the metadata; or if iterable, a tuple of items from the metadata
        :param _all: Returns the entire metadata container
        :return: (key is None) Dictionary of metadata keyed by headers,
          (key is iterable) Tuple of items pulled from metadata,
          (else) item requested or None if not present.
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            data = json.load(fh)
        iteration = str(iteration)
        abscissa = data.get('abscissa', {})
        metadata_container = abscissa.get('metadata', {})
        if _all:
            return metadata_container
        metadata = metadata_container.get(iteration, {})
        if key:
            if isinstance(key, (tuple, list)):
                return tuple(metadata.get(k, None) for k in key)
            else:
                return metadata.get(key, None)
        else:
            return metadata

    def well_ids(self):
        """
        Generator for all the ids for wells in the project.

        :return: [k for k in keys if k is a valid well ID]
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            data = json.load(fh)
        return [k for k in data.keys() if k in WELL_IDS]

    def history_already_contains(self, key):
        self._touch()
        with open(self.filepath, 'r') as fh:
            project: dict = json.load(fh)
        project_history = project.get('history', [])
        return any([key in entry for entry in project_history])

    @via_lock
    def commit_property(self, prop: str, headers, values: dict, log: Union[list, str]) -> dict:
        """
        Adds calculated properties to a project file

        :param prop: The key the data will be stored under
        :param headers: An iterable of headers under which to assign values
        :param values: The value assigned to the key 'prop' (dict)
        :param log: A string or list of string recounting what processing has occurred.
        :return: project['calculated_properties'][prop]
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            project: dict = json.load(fh)

        project.setdefault('calculated_properties', dict())
        project['calculated_properties'].setdefault(prop, dict())
        selection = project['calculated_properties'][prop]
        if headers:
            for header in headers[:-1]:
                selection.setdefault(header, dict())
                selection = selection[header]
            selection.update({headers[-1]: values})
        else:
            project['calculated_properties'][prop] = values

        project.setdefault('history', list())
        if isinstance(log, list):
            project['history'].extend(log)
        if isinstance(log, str):
            project['history'].append(log)

        with open(self.filepath, 'w') as fh:
            json.dump(project, fh)

        return project['calculated_properties'][prop]

    def get_property(self, prop) -> dict:
        """
        Getter for calculated properties

        :param prop: They key being searched for
        :return: project['calculated_properties'][prop]
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            project: dict = json.load(fh)

        project.setdefault('calculated_properties', dict())
        return project['calculated_properties'].get(prop, dict())

    @via_lock
    def pull_project_copy(self):
        """
        Provides the entire project as a dictionary to allow access to all data.  Requires knowledge of the project
        structure (unlike helper methods).

        **1st Level Keys**
          * method_specifications -> Dict repr of :class:`SparkMethod`
          * abscissa -> Dict
          * history -> List
          * `Well ID` -> {str(`iteration #`): Dict repr of well-iteration data}
          * calculated_properties -> Dict

        **Abscissa Keys**
          * x-axis -> {str(`axis key`): [x axis values]}
          * metadata -> {str(`iteration #`): Dict repr of headers from Spark raw data file}

        **Well-Iteration Data**
          * x-axis -> an axis key (used to find the x-axis from metadata)
          * y-axis -> [signal values]
          * y-axis (processed) -> [signal values after background subtraction(s)]
          * well_context -> {A copy of the well context from the characterization plate DB document}

        **Calculated Properties**
          * Keys are determined by the 'prop' and 'headers' arguments of the commit_property() method
          * Values are determined by the dictionary passed as the 'values' argument of the commit_property() method
          * E.g. arguments (prop="LogP", headers=None) --> project['calculated_properties']['LogP']
          * E.g. arguments (prop="Peaks", headers = ["0", ]) --> project['calculated_properties']['Peaks']['0']

        :return: Dictionary representation of the project file
        """
        self._touch()
        with open(self.filepath, 'r') as fh:
            project: dict = json.load(fh)
        return project

    # MUST be last thing in class! (Otherwise the compiler will freak out when making the @via_lock wrapper
    via_lock = staticmethod(via_lock)


class EventLogger:
    """ Creates stamps in a manifest when events occur """
    def __init__(self):
        """ Creates a thread-safe object for posting messages in a manifest (a list)

        """
        self.manifest = list()
        self._lock = Lock()

    def set_manifest(self, manifest):
        self.manifest = manifest

    def stamp_manifest(self, caller: str,
                       message: str,
                       job_id: str = None,
                       timestamp: Union[str, datetime.datetime] = None)\
            -> Tuple[str, str, str, str]:
        """
        Adds an element to the manifest

        :param caller: Name of the caller
        :param message: A message
        :param job_id: A job_id
        :param timestamp: A timestamp
        :return: (timestamp str form, message, job_id, caller's name)
        """
        if timestamp is None:
            timestamp = datetime.datetime.now().strftime(TIME_FORMAT)
        elif isinstance(timestamp, datetime.datetime):
            timestamp = timestamp.strftime(TIME_FORMAT)
        update = (timestamp, message, job_id, caller,)

        if not self._lock.acquire(timeout=15):
            return update
        try:
            self.manifest.append(update)
            return update
        finally:
            self._lock.release()

    def get_job(self, job_id):
        return sorted([item for item in self.manifest if item[2] == job_id],
                      key=lambda x: datetime.datetime.strptime(x[0], TIME_FORMAT))

    def excise_job(self, job_id):
        job_manifest = self.get_job(job_id)
        self.manifest = list_subtract(self.manifest, job_manifest)
        return sorted(job_manifest, key=lambda x: datetime.datetime.strptime(x[0], TIME_FORMAT))

    def log_scan_complete(self, caller, results, job_id, timestamp=None):
        caller = type(caller).__name__
        return self.stamp_manifest(caller, results, job_id, timestamp)

    def stamp(func):
        """ Wrapper function for the lock/try-foo/finally-unlock workflow

        Implements a timeout of 10 seconds for the lock

        :param func: Wrapped function
        :return: Return of Wrapped function
        :raise TimeoutError: On Timeout
        """
        @wraps(func)
        def wrapper_function(self, caller, job_id=None, timestamp=None, notes=None):
            caller = type(caller).__name__
            message: str = func.__name__
            message = message.replace("_", " ").title()
            if notes:
                message += " [Addendum: " + str(notes) + "]"
            if "Job" in message and job_id is None:
                raise TypeError("stamp_manifest() missing 1 contextual argument: 'job_id'")
            return self.stamp_manifest(caller, message, job_id, timestamp)

        return wrapper_function

    @stamp
    def job_start(self, caller, job_id, timestamp=None):
        return

    @stamp
    def job_end(self, caller, job_id, timestamp=None):
        return

    @stamp
    def job_fail(self, caller, job_id, timestamp=None, notes=None):
        return

    @stamp
    def scan_start(self, caller, job_id=None, timestamp=None):
        return

    @stamp
    def scan_end(self, caller, job_id=None, timestamp=None):
        return

    @stamp
    def scan_partial(self, caller, job_id=None, timestamp=None):
        return

    @stamp
    def scan_cancel(self, caller, job_id=None, timestamp=None, notes=None):
        return

    @stamp
    def scan_fail(self, caller, job_id=None, timestamp=None, notes=None):
        return

    @stamp
    def tray_out_left(self, caller, job_id=None, timestamp=None):
        return

    @stamp
    def tray_out_right(self, caller, job_id=None, timestamp=None):
        return

    @stamp
    def tray_in(self, caller, job_id=None, timestamp=None):
        return

    @stamp
    def photoreactor_load(self, caller, job_id=None, timestamp=None):
        return

    @stamp
    def photoreactor_unload(self, caller, job_id=None, timestamp=None):
        return

    @stamp
    def touch(self, caller, job_id=None, timestamp=None, notes=None):
        return

    stamp = staticmethod(stamp)

    SC_FLAG = "$not_success$"


class Appellomancer:
    """ Appellomancer (class: naming wizard)

    A tool to standardize the naming conventions for the files associated with Spark data collection and processing
    within a queue environment.

    Structure: Within the destination path (Default: "C:/Users/admin/Documents/MCN_Spark_Data"), each queue receives a
    folder.  Within each queue's folder, there is a folder for each assay.  Within each assay's sub-folder, there is
    a project file (json) and copies of the raw Spark data (xlsx).  Raw Spark data files sourced from the source path
    (Default: "C:/Users/Public/Documents/Tecan/SparkControl/Workspaces").
    """
    def __init__(self, queue_name, assay, dst_path=_DEFAULT_MCN_DIR, src_path=_DEFAULT_SAVE_DIR):
        """
        Establishes the context for the Appellomancer to a queue and assay.  Allows for the base directories to be
        modified as needed.

        :param queue_name: Name of the queue
        :param assay: Type of the assay (e.g. 'peaks', 'kinetic', etc.)
        :param dst_path: Destination path (where projects should be created)
        :param src_path: Source path (where raw data is deposited from Spark)
        """
        self.queue_name = queue_name
        self.assay_type = assay
        self.dst_path = dst_path
        self.src_path = src_path

    def setter(self, queue_name, assay):
        """
        Setter to allow the context to be changed if needed.

        :param queue_name: Name of the queue
        :param assay: Type of the assay (e.g. 'peaks', 'kinetic', etc.)
        :return: None
        """
        self.queue_name = queue_name
        self.assay_type = assay

    def get_path_to_project(self) -> str:
        """
        Creates a full path to the project file.

        Format: "{dst_path}/queue_{queue_name}/{queue_name}_{assay_type}_project.json"

        :return: The project file path (created using os.path.join)
        """
        return path.join(self.dst_path,
                         f"queue_{self.queue_name}",
                         self.assay_type,
                         f"{self.queue_name}_{self.assay_type}_project.json")

    def get_project_directory(self):
        return self.get_parent_directory(self.get_path_to_project())

    def get_path_to_source_data(self, check=""):
        """
        Locates the raw data produced by the Spark from a scan.

        :param check: [default = "", matches anything] If None, uses the standardized method name
          (see self.generate_method_name method). If str, requires the raw data file contains the 'check' parameter.
        :return: A file path or None
        """
        if check is None:
            check = self.generate_method_name()
        spark_output_path = grab_newest_folder(self.src_path)
        if check not in spark_output_path:
            return None
        spark_data_path = path.join(spark_output_path, r"Export\xlsx")
        try:
            srcnamespace = listdir(spark_data_path)
        except FileNotFoundError:
            return None
        if not srcnamespace:
            return None
        for file in srcnamespace:
            spark_data_full_path = path.join(spark_data_path, file)
            if ".xlsx" in file:
                return spark_data_full_path
        else:
            return None

    def generate_raw_data_dst_name(self, iteration, attempt):
        """
        Used to create standardized names for raw data files.

        Format: "{queue_name}_{iteration, n-digits: 4, L-pad: 0}_{attempt, n-digits: 2, L-pad: 0}.xlsx"

        :param iteration: The iteration counter
        :param attempt: The attempt counter (when an iteration must be retried due to warning/error)
        :return: The file path (created using os.path.join)
        """
        project_path = self.get_project_directory()
        output_file = f"{self.queue_name}_{str(iteration).rjust(4, '0')}_{str(attempt).rjust(2, '0')}.xlsx"
        return path.join(project_path, output_file)

    @staticmethod
    def generate_safe_name(filepath, extension=None):
        """
        Helper method for creating file names without overwriting existing files.

        Mimics Windows default for copy ("file.txt", "file (1).txt", "file (2).txt", etc.)

        :param filepath: The filepath to a desired file.
        :param extension: The file extension being used.
        :return: If file does not already exist, it returns filepath unaltered.  If the file does exist, it returns
          the filepath but with "(#)" appended to the filename, starting at 1.

        :raises FileExistsError: If the counter exceeds 99.
        """
        if not path.exists(filepath):
            return filepath

        if extension:
            bare_path = "".join(filepath.rsplit(extension, 1))
        else:
            bare_path = filepath
            extension = ""
        ii = 1
        while ii < 100:
            new_path = bare_path + f" ({ii})" + extension
            if not path.exists(new_path):
                return new_path
            ii += 1
        raise FileExistsError("Cannot create safe file name")

    def generate_method_name(self):
        """
        Standard form for the naming of a method

        :return: "{queue_name}_{assay_type}"
        """
        return f"{self.queue_name}_{self.assay_type}"

    @staticmethod
    def get_parent_directory(file_path) -> str:
        """
        Helper method to call os.path.split without importing the os module again

        :param file_path: A path-like string
        :return: The first argument from os.path.split, the parent directory of the given path.
        """
        parent, _ = path.split(file_path)
        return parent

    @staticmethod
    def get_filename(file_path) -> Tuple[str, str]:
        """
        Helper method to call os.path.split without importing the os module again

        :param file_path: A path-like string
        :return: The second argument from os.path.split, the file name of the given path.  However, the file name is
          split on "." (from the right) to return a tuple (bare_file_name, extension)
        """
        _, filename = path.split(file_path)
        name, extension = filename.rsplit(".", 1)
        return name, extension

    @staticmethod
    def verify_assay_type(assay):
        """
        Helper method to ensure that an assay name is recognized.

        Supported Assays: 'peaks', 'logd', 'logpi', 'oxideg', 'photodeg', 'hyperpol', and 'tox'

        :param assay: A string representation of the assay name
        :return: True if valid, False if not.
        """
        if assay in ['peaks',
                     'logd', 'logpi;'
                     'oxideg', 'photodeg',
                     'hyperpol',
                     'tox',
                     'check']:
            return True
        return False
