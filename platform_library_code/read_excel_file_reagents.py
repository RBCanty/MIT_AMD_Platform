# -*- coding: utf-8 -*-
"""
Created on Thu Nov  3 17:29:51 2022

@author: admin
"""

import openpyxl
import sys
import re
from pprint import pprint
import json


def search_value_in_col_idx(ws, search_string, col_idx=0):
    values = []
    for row in range(1, ws.max_row + 1):
        if ws[row][col_idx].value == search_string:
            values.append(row)
    return values


def find_bookend_rows(ws):
    row_count = ws.max_row
    begin_lines = search_value_in_col_idx(ws, '<<<<BEGIN>>>>')
    ending_lines = search_value_in_col_idx(ws, '<<<<END>>>>')
    if len(begin_lines) != len(ending_lines):
        print('Unmatched lines for the reimport')
        sys.exit()
    sheetranges = list(zip(begin_lines, ending_lines) )
    return sheetranges


wb = openpyxl.load_workbook(r'.\library_excel_files\current_database.xlsx')
sheetnames = wb.sheetnames

sheetname = 'reagents'
ws = wb[sheetname]
document_ranges = find_bookend_rows(ws)

new_documents = {}
details = 0
contents = 0
update_problems = []
if sheetname not in new_documents.keys():
    new_documents[sheetname] = {}
for document_range in document_ranges:
    document_details = {}
    for line_index in range(*document_range):
        current_cell = ws.cell(line_index, 1).value
        if current_cell is None:
            continue
        if '<<<<BEGIN>>>>' in current_cell or '<<<<END>>>>' in current_cell:
            continue
        if 'DETAILS FOR LABWARE' in current_cell:
            details = 1
            continue
        if 'CONTENTS OF LABWARE' in current_cell:
            details = 0
            contents = 1
            max_column = ws.max_column
            content_keys = [ws.cell(line_index+1, col_index+1).value for col_index in range(0, max_column)
                            if ws.cell(line_index+1, col_index+1).value != '']
            continue
        
        if details == 1:
            detail_type = current_cell
            if detail_type == 'location':
                detail_value = [ws.cell(line_index, 2).value, json.loads(ws.cell(line_index, 3).value)]
            else:
                detail_value = ws.cell(line_index, 2).value
            document_details[detail_type] = detail_value
        elif contents == 1:
            if 'contents' not in document_details.keys():
                document_details['contents'] = {}
            if re.match(r'[A-Z]{1}\d{1,2}', current_cell):
                well_contents = {}
                for content_index, content_key in enumerate(content_keys):
                    cell_contents = ws.cell(line_index, content_index+1).value
                    if cell_contents is not None:
                        well_contents[content_key] = cell_contents
                if 'chemical_smiles' in well_contents.keys() and 'chemical_name' in well_contents.keys():
                    document_details['contents'][current_cell] = well_contents
    if document_details['container_name'] in new_documents.keys():
        print('Duplicated container_name in the %s collection, fix this' % sheetname)
        sys.exit()
    new_documents[sheetname][document_details['container_name']] = document_details


# With the documents re-parsed we need to get them back into the database
import pymongo
from platform_library_functions import query_collection, update_database, \
            query_document, update_database_document, add_to_database
from copy import deepcopy


def without_keys(d, keys):
    return {k: v for k, v in d.items() if k not in keys}


carriers_and_labware_path = r'C:\Users\admin\PycharmProjects\AMD_Control_Platform\venv\Shell_AH\Evoware_API\evoware_library_files\carriers_and_labware.json'
with open(carriers_and_labware_path, 'r') as jsonfile:
    carriers_and_labware = json.load(jsonfile)


platform_mongodb = pymongo.MongoClient(host = '127.0.0.1', port = 27017, 
                                       username = 'XXX', 
                                       password = 'XXX', 
                                       authSource = 'admin')

for collection in new_documents.keys():
    for document in new_documents[collection].keys():
        document_details = new_documents[collection][document]
        if document_details['_id'] == 'NEW_DOCUMENT':
            new_document = without_keys(deepcopy(document_details), ['_id'])
            return_flag, request_return = add_to_database(new_document, 
                                                          collection, 
                                                          carriers_and_labware, 
                                                          platform_mongodb)
            if return_flag != 'Success':
                print('Something failed in adding a new document to database')
                pprint(without_keys(document_details, ['contents']))
                pprint(request_return)
                sys.exit()
            else:
                print(return_flag, request_return)
        else:
            existing_document, response = query_document(collection, 
                                                         document_details['_id'], 
                                                         '_id', platform_mongodb)
            if type(existing_document) != dict:
                print('Something failed when looking up document that should exist: %s' % document_details['_id'])
                sys.exit()
            flag_return, response = update_database_document(existing_document,
                                                             document_details,
                                                             collection,
                                                             platform_mongodb)
            print(flag_return, response)
            if flag_return != 'Success':
                print('Something failed during database updating')
                sys.exit()
            
            
platform_mongodb.close()









