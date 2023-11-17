# -*- coding: utf-8 -*-
"""
Created on Wed Jul 14 11:48:12 2021

@author: Brent
"""
import pymongo, sys
from pprint import pprint
from platform_library_functions import query_collection, update_database, query_document, update_database_document
import datetime
import copy
import json
import natsort

platform_mongodb = pymongo.MongoClient(host = '127.0.0.1', port = 27017, 
                                       username = 'USER', password = 'PASSWORD', 
                                       authSource = 'admin')
    
important_collection = 'reagents'

message = {'auth': {'port': '27017', 'user': 'XXX', 'password': 'XXX'},
           'request':{'request_type':'query_collection', 'collection': important_collection}}
document_return, extra_statements = query_collection(message['request']['collection'], platform_mongodb)
gathered_documents = document_return

platform_mongodb.close()

import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter
import json

carriers_and_labware_path = r'C:\Users\admin\PycharmProjects\AMD_Control_Platform\venv\Shell_AH\Evoware_API\evoware_library_files\carriers_and_labware.json'
with open(carriers_and_labware_path, 'r') as jsonfile:
    carriers_and_labware = json.load(jsonfile)


wb = openpyxl.Workbook()
wb.remove(wb.active)
ws = wb.create_sheet('reagents')
current_line = 1
for current_document in gathered_documents.keys():
    current_details = gathered_documents[current_document]
    # Start by writing the header of the output excel file
    ws.cell(current_line, 1, value='<<<<BEGIN>>>>')
    current_line += 1
    ws.cell(current_line, 1, value='DETAILS FOR LABWARE')
    current_cell = ws.cell(current_line, 1)
    current_cell.font = Font(color='ae0000', bold=True, italic=True)
    current_line += 1
    labels = ['_id', 'container_name', 'location', 'labware_type', 'date_created']
    for label in labels:
        if label == 'location':
            ws.cell(current_line, 1, value=label)
            ws.cell(current_line, 2, value=str(current_details[label][0]))
            ws.cell(current_line, 3, value=str(current_details[label][1]))
        else:
            ws.cell(current_line, 1, value=label)
            ws.cell(current_line, 2, value=str(current_details[label]))
        current_cell = ws.cell(current_line, 1)
        current_cell.font = Font(bold=True)
        current_line += 1
    
    current_line += 1
    ws.cell(current_line, 1, value='CONTENTS OF LABWARE')
    current_cell = ws.cell(current_line, 1)
    current_cell.font = Font(color='ae0000', bold=True, italic=True)
    current_line += 1
    relevant_details = ['plate_well', 'chemical_name', 'chemical_smiles', 'container_name', 
                        'concentration_molar', 'volume_ul', 'solvent']
    for index, detail_name in enumerate(relevant_details):
        ws.cell(current_line, index+1, value=detail_name)
        current_cell = ws.cell(current_line, index+1)
        current_cell.font = Font(color='00008B', bold=True, italic=True)
        ws.column_dimensions[get_column_letter(index + 1)].width = 25
    current_line += 1
    
    # Now populate the contents section of the excel document
    labware_type = current_details['labware_type']
    labware_details = carriers_and_labware['labware'][labware_type]
    allowed_wells = []
    for i in range(0, labware_details['nrows'] * labware_details['ncols']):
        row = i // labware_details['ncols']
        col = i % labware_details['ncols']
        allowed_wells.append(chr(ord('A') + row) + str(col + 1))

    
    well_key_list = natsort.natsorted(list(current_details['contents'].keys()))
    for well_key in allowed_wells:
        if well_key not in well_key_list:
            ws.cell(current_line, 1, value=well_key)
        else:
            for index, detail_name in enumerate(relevant_details):
                if detail_name not in current_details['contents'][well_key].keys():
                    continue
                ws.cell(current_line, index+1, value=current_details['contents'][well_key][detail_name])
        current_line += 1 
    current_line += 1
    ws.cell(current_line, 1, value='<<<<END>>>>')
    current_line += 3

wb.save(r'.\library_excel_files\current_database.xlsx')






























